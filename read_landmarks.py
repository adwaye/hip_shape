from stl import mesh
import os
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from skspatial.objects import Plane, Points
from skspatial.plotting import plot_3d
from mpl_toolkits.mplot3d import Axes3D

# path = './data/UCLH-Controls'
#
# studies = [f for f in os.listdir(path) if os.path.isdir(os.path.join(path,f))]
#
# study = studies[0]
#
#
# study_path = os.path.join(path,study)
# xlsx_file = [f for f in os.listdir(study_path) if f.split('.')[-1]=="xlsx"][0]
#
#
# xlsx_path  = os.path.join(study_path,xlsx_file)
# # Define variable to load the wookbook
# workbook = openpyxl.load_workbook(xlsx_path,data_only=True)
#
# # Define variable to read the active sheet:
# worksheet = workbook.active


# # Iterate the loop to read the cell values
# for i in range(0, worksheet.max_row):
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         print(col[i].value, end="\t\t")
#     print('')


def find_structure_coordinate_socket(worksheet):
    """
    Finds the coordinates of landmarks from an excel file. this assumes that the excel file does not follow any specific
    format and is arranged in a disorderly manner
    :param worksheet: object returned by openpyxl.load_workbook().active
    :return: dictionary of landmarks column and row where these coordinates point to where one should iterate to get
    the landmarks
    """


    # Iterate the loop to read the cell values
    #rows start at 1
    #columns start at 0
    my_dict = {}
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if type(col[i].value) is str:
                if "ant lat" in col[i].value.lower():
                    print("found ant column at "+col[i].column_letter+str(col[i].row))

                    ant_col = col[i].column-1
                    ant_row = i+1
                    if type(worksheet[ant_row-1][ant_col].value) is str:
                        if "right" in worksheet[ant_row-1][ant_col].value.lower():
                            my_dict['Right Ant Lat Col'] = ant_col
                            my_dict['Right Ant Lat Row'] = ant_row
                    if type(worksheet[ant_row-1][ant_col].value) is str:
                        if "left" in worksheet[ant_row-1][ant_col].value.lower():
                            my_dict['Left Ant Lat Col'] = ant_col
                            my_dict['Left Ant Lat Row'] = ant_row
                elif "post lat" in col[i].value.lower():
                    print("found post column at "+col[i].column_letter+str(col[i].row))

                    ant_col = col[i].column-1
                    ant_row = i+1
                    if type(worksheet[ant_row-1][ant_col-3].value) is str:
                        if "right" in worksheet[ant_row-1][ant_col-3].value.lower():
                            print('saving right post lat coords')
                            my_dict['Right Post Lat Col'] = ant_col
                            my_dict['Right Post Lat Row'] = ant_row
                    if type(worksheet[ant_row-1][ant_col-3].value) is str:
                        if "left" in worksheet[ant_row-1][ant_col-3].value.lower():
                            my_dict['Left Post Lat Col'] = ant_col
                            my_dict['Left Post Lat Row'] = ant_row

    return my_dict



def axisEqual3D(ax):
    extents = np.array([getattr(ax, 'get_{}lim'.format(dim))() for dim in 'xyz'])
    sz = extents[:,1] - extents[:,0]
    centers = np.mean(extents, axis=1)
    maxsize = max(abs(sz))
    r = maxsize/2
    for ctr, dim in zip(centers, 'xyz'):
        getattr(ax, 'set_{}lim'.format(dim))(ctr - r, ctr + r)


def find_coordinates_from_worksheet(worksheet,my_dict,key="Right Ant Lat"):
    """
    extracts the socket coordinates from an excelt sheet by using the coordinates where the search should start
    :param worksheet: loaded excel file by pyopenxl
    :param my_dict:   dictionary containing column and row indices for where the information can be found
    :param key: Which side and part of the socket to look for
                choices : "Right Ant Lat", "Right Post Lat", "Left Ant Lat", "Left Post Lat"
                these are keys of the dictionary output by find_structure_coordinate_socket
    :return:
    """
    row_val = my_dict[key+' Row']
    col_val = my_dict[key+' Col']
    ix=1
    jx=1
    coordinates = None

    while worksheet[row_val+ix][col_val].value is not None:
        coord_list = []
        for k in range(0,3):
            coord_list+=[worksheet[row_val+ix][col_val+k].value]
        if ix==1:
            coordinates = np.expand_dims(np.array(coord_list),0)
        else:
            coordinates = np.concatenate((coordinates,np.expand_dims(np.array(coord_list),0)),axis=0)
        ix+=1
    return coordinates

def find_imOriginSpacing_from_worksheet(worksheet):
    """
    Reads the image origin and spacing from the worksheet corresponding to a landmark xlsx file from the landmark data
    :param worksheet: worksheet loaded from openpyxl.load_workbook().active
    :return: list of np array of shape [1,3] showing the location of the image origin and the spacing of the origin
    """
    my_dict = {}
    origin  = []
    spacing = []
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if type(col[i].value) is str:
                if "image origin" in col[i].value.lower():
                    # print("found image origin "+col[i].column_letter+str(col[i].row))

                    col_val = col[i].column

                    for k in range(0,3):
                        origin += [worksheet[i+1][col_val + k].value]
                        spacing += [worksheet[i + 2][col_val + k].value]
                        # print("col val={:}".format(col_val))
                        # print("row val={:}".format(i+1))
                # else:
                #     print("error no image origin information in file")
    origin = np.array(origin)
    origin = np.expand_dims(origin,axis=0)
    spacing = np.array(spacing)
    spacing = np.expand_dims(spacing,axis=0)
    return origin,spacing





def plot_landmark_uclh_controls(path = './data/Segmentation_and_landmarks_raw/UCLH - Controls', k = 10):
    """
    plots the landmark on top of the stl file representing the pelvis surface for the folder UCLH - Controls
    :param k:
    :return:
    """
    #path = './data/Segmentation_and_landmarks_raw/UCLH - Controls' #transformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - DDH' #transformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - FAI'#tranformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - Controls'#transformation works
    #path = './data/Segmentation_and_landmarks_raw/UCLH - Dysplastics' #transformation does not work


    studies = [f for f in sorted(os.listdir(path)) if os.path.isdir(os.path.join(path,f))]

    study = studies[k]





    study_path = os.path.join(path,study)

    xlsx_file = [f for f in os.listdir(study_path) if f.split('.')[-1] == "xlsx"][0]

    xlsx_path = os.path.join(study_path,xlsx_file)

    # Define variable to load the wookbook
    workbook = openpyxl.load_workbook(xlsx_path,data_only=True)

    # Define variable to read the active sheet:
    worksheet = workbook.active
    origin,spacing = find_imOriginSpacing_from_worksheet(worksheet)

    my_dict = find_structure_coordinate_socket(worksheet)
    fig = plt.figure(figsize=(4,4))

    ax = fig.add_subplot(111, projection='3d')

    ax.set_box_aspect(aspect=(1,1,1))


    folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'left' in f.lower()]
    if len(folders)>0:
        study_subpath = os.path.join(study_path,folders[0])
        stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
        stl_path = os.path.join(study_subpath,stl_file)
        mesh_file = mesh.Mesh.from_file(stl_path)
        point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
        ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen',
                   label='Stl surface')
    folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'right' in f.lower()]
    if len(folders)>0:
        study_subpath = os.path.join(study_path,folders[0])
        stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
        stl_path = os.path.join(study_subpath,stl_file)
        point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
        ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')

    print(study)
    for key in ['Right Ant Lat','Right Post Lat','Left Ant Lat','Left Post Lat']:
        coords  = find_coordinates_from_worksheet(worksheet=worksheet,my_dict=my_dict,key=key)
        print(key)
        print(coords)
        if coords is not None:
            mean_coords = np.mean(coords)
            #scaling = np.array([[0.8379,0.8379,1.5]])
            #origin  = np.array([[-209.1,-375.1,-760.5]])
            #coords = ((coords-mean_coords)*spacing)+mean_coords
            coords = coords+(origin)/spacing
            #
            #coords = scaling*coords
            ax.plot(coords[:,0],coords[:,1],coords[:,2],color='blue',label='Socket plane points')
            ax.scatter(coords[:,0],coords[:,1],coords[:,2],color='black',s=2)  # plot the
            # point (2,3,4) on the figure

    # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'left' in f.lower()]
    # study_subpath = os.path.join(study_path,folders[0])
    # stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
    # stl_path = os.path.join(study_subpath,stl_file)
    # point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
    # ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
    # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'right' in f.lower()]
    # study_subpath = os.path.join(study_path,folders[0])
    # stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
    # stl_path = os.path.join(study_subpath,stl_file)
    # point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
    # ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
    #axisEqual3D(ax)
    fig.suptitle("Folder ="+study_path+" k = {:}".format(k))
    return fig, ax


def plot_landmark(path = './data/Segmentation_and_landmarks_raw/UCLH - Controls', k = 10):
    """
    plots the landmark on top of the stl file representing the pelvis surface
    the difference with the plot_landmark_uclh_controls is that we operate in the mm space as opposed to the voxel scale
    :param k:
    :return:
    """
    #path = './data/Segmentation_and_landmarks_raw/UCLH - Controls' #transformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - DDH' #transformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - FAI'#tranformation works
    #path = './data/Segmentation_and_landmarks_raw/TOH - Controls'#transformation works
    #path = './data/Segmentation_and_landmarks_raw/UCLH - Dysplastics' #transformation does not work


    studies = [f for f in sorted(os.listdir(path)) if os.path.isdir(os.path.join(path,f))]

    study = studies[k]





    study_path = os.path.join(path,study)

    xlsx_file = [f for f in os.listdir(study_path) if f.split('.')[-1] == "xlsx"][0]

    xlsx_path = os.path.join(study_path,xlsx_file)

    # Define variable to load the wookbook
    workbook = openpyxl.load_workbook(xlsx_path,data_only=True)

    # Define variable to read the active sheet:
    worksheet = workbook.active
    origin,spacing = find_imOriginSpacing_from_worksheet(worksheet)

    my_dict = find_structure_coordinate_socket(worksheet)
    fig = plt.figure(figsize=(4,4))

    ax = fig.add_subplot(111, projection='3d')

    ax.set_box_aspect(aspect=(1,1,1))


    folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'left' in f.lower()]
    if len(folders)>0:
        study_subpath = os.path.join(study_path,folders[0])
        stl_file_list = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl']
        if len(stl_file_list)>0:
            stl_file = stl_file_list[0]
            stl_path = os.path.join(study_subpath,stl_file)
        else:
            stl_file_list = [f for f in os.listdir(study_path) if f.split('.')[-1] == 'stl']
            stl_path = os.path.join(study_path,stl_file_list[0])
        point_cloud = mesh.Mesh.from_file(stl_path).v0#/spacing
        ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen',
                   label='Stl surface')
    folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'right' in f.lower()]
    if len(folders)>0:
        study_subpath = os.path.join(study_path,folders[0])
        stl_file_list = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl']
        if len(stl_file_list)>0:
            stl_file = stl_file_list[0]
            stl_path = os.path.join(study_subpath,stl_file)
        else:
            stl_file_list = [f for f in os.listdir(study_path) if f.split('.')[-1] == 'stl']
            stl_path = os.path.join(study_path,stl_file_list[0])

        point_cloud = mesh.Mesh.from_file(stl_path).v0#/spacing
        ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')

    print(study)
    iter = 0
    for key in ['Right Ant Lat','Right Post Lat']:
        coords  = find_coordinates_from_worksheet(worksheet=worksheet,my_dict=my_dict,key=key)
        print(key)
        print(coords)
        if coords is not None:
            mean_coords = np.mean(coords)
            #scaling = np.array([[0.8379,0.8379,1.5]])
            #origin  = np.array([[-209.1,-375.1,-760.5]])
            #coords = ((coords-mean_coords)*spacing)+mean_coords
            coords = coords*spacing+(origin)

            #
            #coords = scaling*coords
            ax.plot(coords[:,0],coords[:,1],coords[:,2],color='blue',label='Socket plane points')
            ax.scatter(coords[:,0],coords[:,1],coords[:,2],color='black',s=2)  # plot the
            if iter ==0:
                coords_big = coords
            else:
                coords_big = np.concatenate((coords_big,coords),axis=0)
            iter +=1
    if coords_big is not None:

        right_plane = Plane.best_fit(Points(coords_big))

        lims = np.min(coords_big,axis=0),np.max(coords_big,axis=0)

        #right_plane.plot_3d(ax,lims_x=(lims[0][0],lims[1][0]),lims_y=(lims[0][1],lims[1][1]))
        right_plane.plot_3d(ax,lims_y=(-150,50),lims_x=(-75,50),alpha=0.2)



    iter = 0
    for key in ['Left Ant Lat','Left Post Lat']:
        coords  = find_coordinates_from_worksheet(worksheet=worksheet,my_dict=my_dict,key=key)
        print(key)
        print(coords)
        if coords is not None:
            mean_coords = np.mean(coords)
            #scaling = np.array([[0.8379,0.8379,1.5]])
            #origin  = np.array([[-209.1,-375.1,-760.5]])
            #coords = ((coords-mean_coords)*spacing)+mean_coords
            coords = coords*spacing+(origin)
            #
            #coords = scaling*coords
            ax.plot(coords[:,0],coords[:,1],coords[:,2],color='blue',label='Socket plane points')
            ax.scatter(coords[:,0],coords[:,1],coords[:,2],color='black',s=2)  # plot the
            if iter ==0:
                coords_big = coords
            else:
                coords_big = np.concatenate((coords_big,coords),axis=0)
            iter +=1
    if coords_big is not None:
        left_plane = Plane.best_fit(Points(coords_big))

        lims = np.min(coords_big,axis=0),np.max(coords_big,axis=0)
        left_plane.plot_3d(ax,lims_y=(-150,50),lims_x=(-75,50),alpha=0.2)

        #right_plane.plot_3d(ax,lims_x=(lims[0][0],lims[1][0]),lims_y=(lims[0][1],lims[1][1]))

        # point (2,3,4) on the figure

    # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'left' in f.lower()]
    # study_subpath = os.path.join(study_path,folders[0])
    # stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
    # stl_path = os.path.join(study_subpath,stl_file)
    # point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
    # ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
    # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'right' in f.lower()]
    # study_subpath = os.path.join(study_path,folders[0])
    # stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
    # stl_path = os.path.join(study_subpath,stl_file)
    # point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
    # ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
    #axisEqual3D(ax)
    fig.suptitle("Folder ="+study_path+" k = {:}".format(k))
    return fig, ax





if __name__=='__main__':
    path = './data/Segmentation_and_landmarks_raw/UCLH - Controls' #transformation works
    studies = [f for f in sorted(os.listdir(path)) if os.path.isdir(os.path.join(path,f))]
    np.random.seed(1234
                   )
    choices = np.random.randint(low=0,high=len(studies),size=1)
    for k in choices.tolist():
        fig,ax = plot_landmark(path=path,k=k)
        study = studies[k]
        study_path = os.path.join(path,study)

        xlsx_file = [f for f in os.listdir(study_path) if f.split('.')[-1] == "xlsx"][0]

        xlsx_path = os.path.join(study_path,xlsx_file)

        # Define variable to load the wookbook
        workbook = openpyxl.load_workbook(xlsx_path,data_only=True)

        # Define variable to read the active sheet:
        worksheet = workbook.active


        my_dict = {}
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if type(col[i].value) is str:
                    if "acetabulum" in col[i].value.lower():
                        print("found column at "+col[i].column_letter+str(col[i].row))

                        ant_col = col[i].column-1
                        ant_row = i
                        for k in range(1,6):
                            if type(worksheet[ant_row+k][ant_col].value) is str:

                                if "rasis" in worksheet[ant_row+k ][ant_col].value.lower():
                                    my_dict['RASIS Col'] = ant_col
                                    my_dict['RASIS Row'] = ant_row+k
                                elif "lasis" in worksheet[ant_row+k ][ant_col ].value.lower():
                                    my_dict['LASIS Col'] = ant_col
                                    my_dict['LASIS Row'] = ant_row+k
                                elif "ltub" in worksheet[ant_row+k ][ant_col ].value.lower():
                                    my_dict['LTUB Col'] = ant_col
                                    my_dict['LTUB Row'] = ant_row+k
                                elif "rtub" in worksheet[ant_row+k][ant_col ].value.lower():
                                    my_dict['RTUB Col'] = ant_col
                                    my_dict['RTUB Row'] = ant_row+k


        it = 0
        for key in ['RASIS','LASIS','RTUB','LTUB']:
            row_val = my_dict[key + ' Row']
            col_val = my_dict[key + ' Col']
            coord_temp = []
            for k in range(1,4):
                if worksheet[row_val][col_val+k].value is not None:
                    coord_temp +=[worksheet[row_val][col_val+k].value]
            if it==0:
                coords = np.array([coord_temp])
            else:
                coords_ = np.array([coord_temp])
                coords  = np.concatenate((coords,coords_),axis=0)

            it+=1


        # coords = np.array([[79.86,-51.64,1141],
        #                    [    -105.1, - 61.88,1141],
        #                    [-0.140000000000015, - 50.36,1050],
        #                    [-22.54, - 51.64,    1048],
        #                    ])
        origin,spacing = find_imOriginSpacing_from_worksheet(worksheet)

        #coords = coords
        ax.scatter(coords[:,0],coords[:,1],coords[:,2],color='red',label='APP points')
        axisEqual3D(ax)
        plt.legend()


        #plot_landmark_uclh_controls(k=7)
        #=========================================================================
        #Attempting to get the right transformation for UCLH - Dysplastics
        #=========================================================================
        # path = './data/Segmentation_and_landmarks_raw/UCLH - Dysplastics'
        #
        # studies = [f for f in sorted(os.listdir(path)) if os.path.isdir(os.path.join(path,f))]
        # k =10
        # study = studies[k]
        #
        #
        #
        #
        # study_path = os.path.join(path,study)
        #
        # xlsx_file = [f for f in os.listdir(study_path) if f.split('.')[-1] == "xlsx"][0]
        #
        # xlsx_path = os.path.join(study_path,xlsx_file)
        #
        # # Define variable to load the wookbook
        # workbook = openpyxl.load_workbook(xlsx_path,data_only=True)
        #
        # # Define variable to read the active sheet:
        # worksheet = workbook.active
        # origin,spacing = find_imOriginSpacing_from_worksheet(worksheet)
        #
        # my_dict = find_structure_coordinate_socket(worksheet)
        # fig = plt.figure(figsize=(4,4))
        #
        # ax = fig.add_subplot(111, projection='3d')
        # ax.set_box_aspect(aspect=(1,1,1))
        #
        #
        # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'left' in f.lower()]
        # if len(folders)>0:
        #     study_subpath = os.path.join(study_path,folders[0])
        #     stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
        #     stl_path = os.path.join(study_subpath,stl_file)
        #     point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
        #     ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
        # folders = [f for f in os.listdir(study_path) if os.path.isdir(os.path.join(study_path,f)) and 'right' in f.lower()]
        # if len(folders)>0:
        #     study_subpath = os.path.join(study_path,folders[0])
        #     stl_file = [f for f in os.listdir(study_subpath) if f.split('.')[-1] == 'stl'][0]
        #     stl_path = os.path.join(study_subpath,stl_file)
        #     point_cloud = mesh.Mesh.from_file(stl_path).v0/spacing
        #     ax.scatter(point_cloud[::10,0],point_cloud[::10,1],point_cloud[::10,2],alpha=0.3,s=1,color='lightgreen')
        #
        #
        # print(study)
        # for key in ['Right Ant Lat','Right Post Lat','Left Ant Lat','Left Post Lat']:
        #     coords  = find_coordinates_from_worksheet(worksheet=worksheet,my_dict=my_dict,key=key)
        #     print(key)
        #     if coords is not None:
        #         print(coords)
        #
        #
        #         mean_coords = np.mean(coords,axis=0,keepdims=True)
        #         #scaling = np.array([[0.8379,0.8379,1.5]])
        #         #origin  = np.array([[-209.1,-375.1,-760.5]])
        #         #coords = ((coords-mean_coords)*spacing)+mean_coords
        #         coords = coords + (origin) / spacing
        #         # mean_coords = np.mean(coords,axis=0,keepdims=True)
        #         # centre_points_ = np.mean(np.concatenate((mean_coords,np.mean(point_cloud,axis=0,keepdims=True))),axis=0)
        #         # centre_points = np.array([[0,0,centre_points_[2]]])
        #         coords = coords-centre_points
        #         coords = coords-2*np.array([[0,0,1]])*coords#reflection in z axis
        #         coords = coords+centre_points
        #
        #         # centre_points = np.array([[0,centre_points_[2],0]])
        #         # coords = coords-centre_points
        #         # coords = coords-2*np.array([[0,1,0]])*coords#reflection in y axis
        #         # coords = coords+centre_points
        #         # #
        #         #coords = scaling*coords
        #         ax.plot(coords[:,0],coords[:,1],coords[:,2],color='blue')
        #         ax.scatter(coords[:,0],coords[:,1],coords[:,2],color='black',s=2)  # plot the point (2,3,4) on the figure
        #
        # axisEqual3D(ax)
        # fig.suptitle("Folder =" + study_path + " k = {:}".format(k))